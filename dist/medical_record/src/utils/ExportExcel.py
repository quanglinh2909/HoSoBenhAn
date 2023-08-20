import openpyxl


def getAddress(member):
    address = ""
    if member.Address != None and member.Address != "":
        address += member.Address
    if member.Ward != None and member.Ward != "":
        address += ", " + member.Ward
    if member.District != None and member.District != "":
        address += ", " + member.District
    if member.Province != None and member.Province != "":
        address += ", " + member.Province
    return address


def create_excel_file(file_path, listMember):
    # Tạo một workbook mới
    workbook = openpyxl.Workbook()

    # Chọn sheet đầu tiên (sheet active mặc định)
    sheet = workbook.active

    # Ghi dữ liệu vào sheet

    sheet['A1'] = 'STT'
    sheet['B1'] = 'Tên bệnh nhân'
    sheet['C1'] = 'Ngày sinh'
    sheet['D1'] = 'CCCD'
    sheet['E1'] = 'Ngày vào'
    sheet['F1'] = 'Tên người nhà'
    sheet['G1'] = 'Thông tin người nhà'
    sheet['H1'] = 'Địa chỉ'

    # set width
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 50
    sheet.column_dimensions['H'].width = 100

    # to mau header
    sheet['A1'].fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    sheet['B1'].fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    sheet['C1'].fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    sheet['D1'].fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    sheet['E1'].fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    sheet['F1'].fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    sheet['G1'].fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    sheet['H1'].fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')


    data = []
    for i in range(len(listMember)):
        data.append((i + 1, listMember[i].FullName, listMember[i].Birthday, listMember[i].CCCD,
                     listMember[i].DateIn,
                     listMember[i].Relatives, listMember[i].InfoRelatives, getAddress(listMember[i])))


    for row_data in data:
        sheet.append(row_data)

    # Lưu workbook xuống tệp Excel
    workbook.save(file_path)
